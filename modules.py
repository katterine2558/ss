from dotenv import load_dotenv
import os
import pyodbc
import requests
import re
import xlwings as wx
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#Carga las variables de entorno
load_dotenv()
API_KEY = os.getenv("CLOCKIFY_API_KEY")
WORKSPACE_ID = os.getenv("CLOCKIFY_WORKSPACE_ID")
SERVER = os.getenv("SERVER")
DATABASE = os.getenv("DATABASE")
USER_ADMIN = os.getenv("USER_ADMIN")
PASSWORD = os.getenv("PASSWORD")
DRIVER = os.getenv("DRIVER")
PUERTO = os.getenv("PUERTO")
ENGINE = os.getenv("ENGINE")

"""
Obtiene la base de datos de los empleados
"""
def get_employees():

    #cadena de conexión SQL Server
    connection_string = (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={SERVER},{PUERTO};"
        f"DATABASE={DATABASE};"
        f"UID={USER_ADMIN};"
        f"PWD={PASSWORD};"
        f"Encrypt=yes;"
        f"TrustServerCertificate=no;"
        f"Connection Timeout=90;"
    )

    try:
        conn = pyodbc.connect(connection_string)
        cursor = conn.cursor()

        query = """
            SELECT 
                ID,
                Email
            FROM 
                RRHH.Colaboradores
        """
        cursor.execute(query)

        # Obtener columnas
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return results  

    except Exception as e:
        raise RuntimeError(f"[ERROR] Error conectando o consultando la base de datos: {e}")

"""
Obtiene los proyectos de clockify
"""
def get_projects():

    # URL para obtener los proyectos del workspace
    url = f'https://api.clockify.me/api/v1/workspaces/{WORKSPACE_ID}/projects'
    headers = {
        "Content-Type": "application/json",
        "x-api-key": API_KEY
    }

    proyectos = []
    page = 1
    try:
        while True:
            payload = {
                "page-size": 200,
                "page": page,
                "archived": False
            }

            response = requests.get(url, headers=headers, params=payload)

            if response.status_code != 200:
                raise RuntimeError(f"[ERROR] Error al obtener proyectos de Clockify en la página {page}: {response.text}")

            data = response.json()

            if not data:
                break  # No hay más proyectos

            proyectos.extend([
                {
                    "id": p["id"],
                    "name": p["name"]
                } for p in data
            ])
            page += 1

        return proyectos

    except Exception as e:
        raise RuntimeError(f"[ERROR] No se pudieron obtener los proyectos: {e}")
    

"""
Obtiene los registros de clockify
"""
def get_records(fecha_inicio, fecha_fin):
    
    url = f"https://reports.api.clockify.me/v1/workspaces/{WORKSPACE_ID}/reports/detailed"
    
    headers = {
        "Content-Type": "application/json",
        "x-api-key": API_KEY
    }

    all_timeentries = []
    page = 1

    try:
        while True:
            payload = {
                "dateRangeStart": f"{fecha_inicio}T00:00:00.000Z",
                "dateRangeEnd": f"{fecha_fin}T23:59:59.999Z",
                "detailedFilter": {
                    "page": page,
                    "pageSize": 1000
                },
                "exportType": "JSON",
                "users": {
                    "status": "ACTIVE"
                }
            }

            response = requests.post(url, headers=headers, json=payload)

            if response.status_code != 200:
                raise RuntimeError(f"[ERROR] Error al obtener los registros de Clockify en la página {page}: {response.text}")

            data = response.json()
            timeentries = data.get("timeentries", [])

            if not timeentries:
                break  # Ya no hay más páginas

            all_timeentries.extend(timeentries)
            page += 1
        
        return all_timeentries

    except Exception as e:
        raise RuntimeError(f"[ERROR] No se pudieron obtener los registros: {e}")
    
"""
Genera el prorrateo de los registros
"""
def get_report(empleados,proyectos, registros,fecha_inicio, fecha_fin):

    #Cambia el formato de fecha de YYYY-MM-DD a DD/MM/YYYY
    fecha_inicio = f"{fecha_inicio.split('-')[2]}-{fecha_inicio.split('-')[1]}-{fecha_inicio.split('-')[0]}"
    fecha_fin =  f"{fecha_fin.split('-')[2]}-{fecha_fin.split('-')[1]}-{fecha_fin.split('-')[0]}"

    dataframe = {
        "Id colaborador": [],
        "Proyecto Id": [],
        "Fecha inicio": [],
        "Fecha fin": [],
        "Prorrateo": []
    }

    # Lista única de correos de los colaboradores
    unique_emails = list({record['userEmail'] for record in registros})

    #Itera por cada correo unico para ver las actividades desarrolladas por ese usuario
    for email in unique_emails:
        try:
            #Identificación del usuario
            empleado = next((e for e in empleados if e["Email"] == email), None)
            if not empleado:
                continue  # salta al siguiente correo si no se encuentra el colaborador
            id_Colaborador = empleado['ID']

            # Registros de un usuario en particular
            user_records = [record for record in registros if record['userEmail'] == email]

            #Proyectos en los que ha trabajado el colaborador
            unique_projects = list({record['projectName'] for record in user_records})

            #T.TRANS lo arregla por TTRANSF
            if "T.TRANS" in unique_projects:
                unique_projects = ["TTRANSF" if item == "T.TRANS" else item for item in unique_projects]
            elif "CAPACITACIONES TÉCNICAS" in unique_projects:
                unique_projects = ["CAPACITACIONES TÉCNICAS" if item == "CAPACITACIONES TECNI" else item for item in unique_projects]

            #Borra los proyectos que no se tienen en cuenta
            proyectos_a_ignorar = {
                "INCAPACIDAD MÉDICA", "COMPENSATORIO", "FESTIVO", "VACACIONES",
                "LICENCIAS MATERNIDAD O PATERNIDAD", "PUENTES", "REHABILITACIONES",
                "INFRAS", "EDIF", "PERMISO AUTORIZADO", "LICENCIA POR FUERZA MAYOR"
            }
            unique_projects = [p for p in unique_projects if p not in proyectos_a_ignorar]

            #Verifica los registros de los proyectos que tienen el patrón P-202X-XX.XX
            user_records, unique_projects = delete_point_proyects(user_records,proyectos,unique_projects)

            #Itera sobre los proyectos para sumar las horas
            horas = []
            for project in unique_projects:

                #Registros asociados a un único proyecto
                project_records = [record for record in user_records if record['projectName'] == project]

                sum_horas=0
                for r in project_records:
                    sum_horas+=(r["timeInterval"]["duration"] / 3600)

                horas.append(sum_horas)

                #Almacena la identificación del usuario
                dataframe["Id colaborador"].append(id_Colaborador)
                #Almacena el identificador del proyecto
                if project.startswith("PM-"): #Proyectos de méxico
                    coincidencias = re.findall(r'\bPM-\d{4}-\d{2}', project)
                    coincidencias[0] = coincidencias[0].split('-')[0]+ '-' + coincidencias[0].split('-')[1][2:] + '-' + coincidencias[0].split('-')[-1]
                elif project.startswith("C") and project[1].isdigit(): 
                    coincidencias = re.findall(r'\bC\d{6}[A-Z]', project)
                elif  project.startswith("P") and project[1].isdigit():
                    coincidencias = re.findall(r'\bP\d{4}-\d{2}\.\d{2}', project)
                elif project.startswith("M") and project[1].isdigit():
                    coincidencias = re.findall(r'\bM\d{4}-\d{2}\.\d{2}', project)
                elif project.startswith("J") and project[1].isdigit():
                    coincidencias = re.findall(r'\bJ\d{4}-\d{2}\.\d{2}', project)
                elif project.startswith("R") and project[1].isdigit():
                    coincidencias = re.findall(r'\bR\d{4}-\d{2}\.\d{2}', project)
                else:
                    coincidencias = re.findall(r'\d{4}-\d{2}', project)

                if len(coincidencias) >0:
                    dataframe["Proyecto Id"].append(coincidencias[0])
                else:
                    dataframe["Proyecto Id"].append(project)
                    
                #Almacena la fecha de inicio y fin
                dataframe["Fecha inicio"].append(fecha_inicio)
                dataframe["Fecha fin"].append(fecha_fin)

            #Ponderacion
            ponderaciones = ponderar_a_100(horas)
            dataframe["Prorrateo"].extend(ponderaciones)

        except Exception as e:
            raise RuntimeError(f"[ERROR] No se pudo realizar el prorrateo {e}")
        
    return dataframe

"""
Escribe el reporte de excel
"""
def write_report(reporte):
    try:
        wb = Workbook()
        ws = wb.active

        # Renombra la hoja con el nombre del prorrateo
        nombre_hoja = reporte["Nombre Prorrateo"][0]
        ws.title = nombre_hoja

        # Escribir los datos fila por fila
        for i in range(len(reporte["Id colaborador"])):
            fila = i + 1

            ws.cell(row=fila, column=1, value=reporte["Id colaborador"][i])
            ws.cell(row=fila, column=2, value=reporte["Proyecto Id"][i])
            ws.cell(row=fila, column=3, value=reporte["Fecha inicio"][i])
            ws.cell(row=fila, column=4, value=reporte["Fecha fin"][i])
            ws.cell(row=fila, column=5, value=reporte["Prorrateo"][i])
            ws.cell(row=fila, column=6, value=reporte["Nombre Prorrateo"][i])

        # Ajustar ancho de columnas
        column_widths = [15, 15, 10, 10, 5, 20]
        for i, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # Ajustar altura de filas
        for fila in range(1, len(reporte["Id colaborador"]) + 1):
            ws.row_dimensions[fila].height = 15

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"{reporte['Nombre Prorrateo'][0]}.xlsx"

        return output, nombre_archivo

    except Exception as e:
        raise RuntimeError(f"[ERROR] No se pudo escribir el reporte: {e}")


#Elimina los proyectos que tienen un patron de 202X-XX.XX en el nombre y asigna los registros de este como si fuese el principal. 
def delete_point_proyects(user_records,projects,unique_projects):
    
    # Patrón para 202X-XX-XX.XX donde X es cualquier número
    PATRON_PREFIJO = r'^202\d-\d{2}\.\d{2}'

    #Itera por los proyectos
    for proj in unique_projects:
        if re.match(PATRON_PREFIJO, proj):
            #Elimina el proyecto en la lista de proyectos 
            unique_projects.remove(proj)
            #Nombre del proyecto sin .XX
            project = [pos for pos in projects if pos["name"].startswith(proj[0:7] + " ")][0]["name"]
            #Itera por los registros de los usuarios
            for r in user_records:
                if r["projectName"] == proj:
                    r["projectName"] = project
    
    return user_records, unique_projects

#Hace la ponderación de horas
def ponderar_a_100(valores):
    suma_total = sum(valores)
    
    if suma_total == 0:
        return [0] * len(valores)  # Para evitar la división por cero
    
    factor_escala = 100 / suma_total
    
    ponderaciones = [int(round(valor * factor_escala)) for valor in valores]
    
    return ponderaciones